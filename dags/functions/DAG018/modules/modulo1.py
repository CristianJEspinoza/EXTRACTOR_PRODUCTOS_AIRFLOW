## Modulo 1 : Verificación de todos los productos emitidos existentes
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import pandas as pd
import os
from openpyxl.styles import PatternFill
# Importando Variables de Entornos
import os


class Modulo1:
    
    api = None
    
    def __init__(self, apiInicio):
        self.apiInicio = apiInicio
        pass
    
    def Start(self):
        self.api = self.apiInicio
        
        print("-- Iniciando proceso de verificación MODULO 1")
        ##
        bar_code = pd.DataFrame({})
        unit_eq = pd.DataFrame({})
        price = pd.DataFrame({})
        provider = pd.DataFrame({})
        traduc1 = pd.DataFrame({})
        traduc2 = pd.DataFrame({})
        traduc3 = pd.DataFrame({})
        des_nom_product = pd.DataFrame({})
        unit_symbol = pd.DataFrame({})
        product_type = pd.DataFrame({})
        product_subtype = pd.DataFrame({})
        service_type = pd.DataFrame({})
        name_fields = pd.DataFrame({})
        storage_dimension = pd.DataFrame({})
        dimension_group_name = pd.DataFrame({})
        item_model = pd.DataFrame({})
        purchase_unit_symbol = pd.DataFrame({})
        sales_unit_symbol = pd.DataFrame({})
        purchase_sales_tax_item = pd.DataFrame({})
        sales_sales_tax_item = pd.DataFrame({})
        purchase_price_automaticall = pd.DataFrame({})
        sales_price_adjustment = pd.DataFrame({})
        net_product = pd.DataFrame({})
        order_underde = pd.DataFrame({})
        inventory_unit_symbol = pd.DataFrame({})
        production_type = pd.DataFrame({})
        product_coverage = pd.DataFrame({})
        product_group_id = pd.DataFrame({})
        unit_cost_automatically = pd.DataFrame({})
        key_in_quantity = pd.DataFrame({})
        is_manual_discount = pd.DataFrame({})
        code_product = pd.DataFrame({})
        code_productEightDigits = pd.DataFrame({})
        inactiveProduct = pd.DataFrame({})

        ###
        tipo_verificacion = []
        
        ## Verificando codigo de barras
        print("-- Verificando codigo de barras")
        try:
            bar_code = self.api.verifyExistsBarCode()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los codigos de barra")
        
        ##Verificando codigo de barras doble
        try:
            dublicate_bar_code = self.api.verifyExistsDuplicateBarCode()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de traducciones2")
        
        ## Verificando unidades de conversion - equivalencias
        print("-- Verificando codigo unidades de conversión")
        try:
            unit_eq = self.api.verifyEquiUnitSymbol()
        except:
            print("ERROR: Ocurrió un error al momento de verificar la unidades de conversion - equivalencias")
        
        ## Verificando asignación de precios
        print("-- Verificando precios de productos")
        try:
            price = self.api.verifyPricesExists()
        except:
            print("ERROR: Ocurrió un error al momento de verificar precios asignados")
            
        ## Verificando codigo proveedor
        print("-- Verificando asignacion de proveedores")
        try:
            provider = self.api.verifyAssignProvider()
        except:
            print("ERROR: Ocurrió un error al momento de verificar asignacion de proveedores")
        ## Verificando traducciones
        print("-- Verificando campos de traducciones")
        try:
            traduc1 = self.api.verifyTranslationProduct1()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de traducciones1")
        try:
            traduc2 = self.api.verifyTranslationProduct2()
        except Exception as e:
            print("ERROR: Ocurrió un error al momento de verificar los campos de traducciones2")
            print(e)
        try:
            traduc3 = self.api.verifyTranslationProduct3()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de traduc3")
        
        #Verificación de productos inactivos
        print("-- Verificando productos inactivos")
        try:
            inactiveProduct = self.api.verifyProductsInStateActive()
        except:
            print("ERROR: Ocurrió un error al momento de verificar asignacion de proveedores")
        
        ##Verificación de descripción y nombre del producto en Traducción 
        try:
            des_nom_product = self.api.verifyTranslationProductosEmptyFields()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de des_nom_product")

        ##Verificación de conversiones de unidades que tengan en Desde unidad U y en Hasta unidad U.
        try:
            unit_symbol = self.api.verifyFromToUnitSymbol()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de unit_symbol")
        
        #Verificar que tipo de producto sea artículo
        try:
            product_type = self.api.verifyProductType()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de product_type")
        
        #verificar que subtipo de producto sea Producto
        try:
            product_subtype = self.api.verifyProductSubType()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de product_subtype")

        #Verificar que el tipo de servicio de producto sea "No especificado"
        try:
            service_type = self.api.verifyServiceType()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de service_type")
        
        #Verificar que los campos de nombre del producto, nombre de busqueca y descripción no esten vacios
        try:
            name_fields = self.api.verifyNameFields()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de name_fields")
        
        #Verificar que los campos de grupo de dimensiones de almacenamiento sea SiteWHL
        try:
            storage_dimension = self.api.verifyStorageDimensionGroupName()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de storage_dimension")
        
        #Verificar que los campos de grupo de dimensiones de seguimiento sea Ninguno
        try:
            dimension_group_name = self.api.verifyTrackingDimensionGroupName()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de dimension_group_name")
        
        #Verificar que los campos de grupo de modelos de artículo sea PROD
        try:
            item_model = self.api.verifyItemModelGroupId()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de item_model")
        
        #Verificar que la unidad del pedido de compra no sea igual a U.
        try:
            purchase_unit_symbol = self.api.verifyPurchaseUnitSymbol()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de purchase_unit_symbol")
        
        #Verificar que la unidad del pedido de ventas no sea igual a U.
        try:
            sales_unit_symbol = self.api.verifySalesUnitSymbol()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de sales_unit_symbol")
        
        #Verificar que el grupo de impuestos de artículos en compras no este vacio
        try:
            purchase_sales_tax_item = self.api.verifyPurchaseSalesTaxItemGroupCode()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de purchase_sales_tax_item")
        
        #Verificar que el grupo de impuestos de artículos en ventas no este vacio
        try:
            sales_sales_tax_item = self.api.verifySalesSalesTaxItemGroupCode()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de sales_sales_tax_item")
        
        #Verificar que el el último precio de compra en actualización del precio en Compra sea Sí
        try:
            purchase_price_automaticall = self.api.verifyIsPurchasePriceAutomaticallyUpdated()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de purchase_price_automaticall")
        
        #Verificar que permitir ajuste de precios en Vender sea No
        try:
            sales_price_adjustment = self.api.verifyIsSalesPriceAdjustmentAllowed()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de sales_price_adjustment")
        
        #Verificar que todos los productos tengan peso neto
        try:
            net_product = self.api.verifyNetProductWeight()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de net_product")
        
        #Verificar campo de entrega incompleta en transfererencia de pedidos de medidas de peso sea igual a 100 
        try:
            order_underde = self.api.verifyTransferOrderUnderdeliveryPercentage()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de order_underde")
        
        #Verificar que le unidad de inventario en administrar inventario sea U.
        try:
            inventory_unit_symbol = self.api.verifyInventoryUnitSymbol()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de inventory_unit_symbol")
        
        #Verificar que tipo de producción en planificación de formula en aplicar ingeniería tiene que ser ninguno:
        try:
            production_type = self.api.verifyProductionType()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de production_type")
        
        #Verificar que el grupo de cobertura de Plan sea Grupo:
        try:
            product_coverage = self.api.verifyProductCoverageGroupId()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de product_coverage")
        
        #Verificar que el grupo de artículos sea GA001
        try:
            product_group_id = self.api.verifyProductGroupId()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de product_group_id")
        
        #Verificar que ultimo precio de coste en gestionar costes este marcado como Sí
        try:
            unit_cost_automatically = self.api.verifyIsUnitCostAutomaticallyUpdated()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de unit_cost_automatically")
        
        #Verificar que especificar precio en comercio tiene que ser iguar a "No se debe especificar el precio"
        try:
            key_in_quantity = self.api.verifyKeyInQuantityRequirementsAtPOSRegister()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de key_in_quantity")
        
        #Verificar que evitar los descuentos manuales en comercio sea igual a Sí
        try:
            is_manual_discount = self.api.verifyIsManualDiscountPOSRegistrationProhibited()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de is_manual_discount")
        
        #Verificar que el código iternacional sea igual al código SUNAT en datos Perú
        try:
            code_product = self.api.verifyCodeProduct()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de code_product")
        
        #Verificar que el código internacional y el código SUNAT tengan 8 dígitos en datos Perú:
        try:
            code_productEightDigits = self.api.verifyCodeProductWithEightDigits()
        except:
            print("ERROR: Ocurrió un error al momento de verificar los campos de code_productEightDigits")
        ##
        ####
        seed_cell_excel = 10
        
        
        # Obtener la ruta al directorio 'assets' desde 'modulo1.py'
        ruta_assets = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'assets'))

        # Construir la ruta al archivo Excel dentro de la carpeta 'assets'
        ruta_archivo = os.path.join(ruta_assets, 'model_reporte.xlsx')

        
        
        data_sheet = openpyxl.load_workbook(ruta_archivo)
        sheet = data_sheet.active
        sheet['C5']= datetime.now()
        ####
        colum_l=['D','E']
        count_f = 0
        ####
        ##PRIMARY
        ####
        if bar_code.empty != True:
            ##
            tipo_verificacion.append("BARCODE")
            ##
            rows = dataframe_to_rows(bar_code,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "ERROR EN CÓDIGO DE BARRAS"
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ####
        if dublicate_bar_code.empty != True:
            ##
            tipo_verificacion.append("DUPLICATEBARCODE")
            ##
            rows = dataframe_to_rows(dublicate_bar_code,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "SE ENCONTRO EL CÓDIGO DE BARRAS REPETIDO."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        if unit_eq.empty != True:
            ##
            tipo_verificacion.append("UNIT")
            ##
            rows = dataframe_to_rows(unit_eq,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        if count_c ==2:
                            sheet[f'C{num_row}'] = f"ERROR EN UNIDADES DE CONVERSIÓN - {str(value)}"
                        else:
                            sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1 
        ####
        ####
        if inactiveProduct.empty != True:
            ##
            tipo_verificacion.append("STATUS_ACTIVE")
            ##
            rows = dataframe_to_rows(inactiveProduct,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "SE VENDIO PRODUCTOS CON ESTADO INACTIVO."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1 
        ##
        ####
        if unit_symbol.empty != True:
            ##
            tipo_verificacion.append("UNITSYMBOL")
            ##
            rows = dataframe_to_rows(unit_symbol,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "NO SE ENCONTRO 'U' EN DESDE UNIDAD Y 'U.' EN HASTA UNIDAD EN LA CONVERSIÓN DE UNIDADES."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1 
        ####
        ##
        if sales_unit_symbol.empty != True:
            ##
            tipo_verificacion.append("SALESUNITSYMBOL")
            ##
            rows = dataframe_to_rows(sales_unit_symbol,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "LA 'UNIDAD' EN PEDIDO DE VENTAS ES IGUAL A 'U.'."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if sales_price_adjustment.empty != True:
            ##
            tipo_verificacion.append("SALESPRICEADJUSTMENT")
            ##
            rows = dataframe_to_rows(sales_price_adjustment,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "'PERMITIR AJUSTES DE PRECIOS' EN 'VENDER' ESTA MARCADO COMO 'SÍ'."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if key_in_quantity.empty != True:
            ##
            tipo_verificacion.append("KEYINQUANTITY")
            ##
            rows = dataframe_to_rows(key_in_quantity,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "EL CAMPO DE 'ESPECIFICAR PRECIO' EN 'COMERCIO' ES DISTINTO A 'No se debe especificar el precio'."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if is_manual_discount.empty != True:
            ##
            tipo_verificacion.append("ISMANUALDISCOUNT")
            ##
            rows = dataframe_to_rows(is_manual_discount,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "'EVITAR LOS DESCUENTOS MANUALES' EN 'COMERCIO' ESTÁ MARCADO COMO 'NO'."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if inventory_unit_symbol.empty != True:
            ##
            tipo_verificacion.append("INVENTORYUNITSYMBOL")
            ##
            rows = dataframe_to_rows(inventory_unit_symbol,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "LA 'UNIDAD' EN 'ADMINISTRAR INVENTARIO' ES DISTINTO A 'U.'."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        # Agregar una celda pintada de amarillo al final del bloque (fuera del bloque if)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        last_row = seed_cell_excel + count_f
        sheet[f'B{last_row}'].fill = yellow_fill
        sheet[f'C{last_row}'].fill = yellow_fill
        sheet[f'D{last_row}'].fill = yellow_fill
        sheet[f'E{last_row}'].fill = yellow_fill
        count_f = count_f+1
        ####
        ##SECONDARY
        ####
        if price.empty != True:
            ##
            tipo_verificacion.append("PRICE")
            ##
            rows = dataframe_to_rows(price,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        if count_c ==2:
                            sheet[f'C{num_row}'] = f"ERROR EN ASIGNACIÓN DE PRECIOS - {str(value)}"
                        else:
                            sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ####
        if provider.empty != True:
            ##
            tipo_verificacion.append("PROVIDER")
            ##
            rows = dataframe_to_rows(provider,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "NO TIENE ASIGNADO PROVEEDOR"
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1 
        ####
        if traduc1.empty != True:
            ##
            if "TRANSLATION" not in tipo_verificacion:
                tipo_verificacion.append("TRANSLATION")
            ##
            rows = dataframe_to_rows(traduc1,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = 'SOLO TIENE TRADUCCIÓN PARA IDIOMA "ES-MX"'
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1 
        ####
        if traduc2.empty != True:
            ##
            if "TRANSLATION" not in tipo_verificacion:
                tipo_verificacion.append("TRANSLATION")
            ##
            rows = dataframe_to_rows(traduc2,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = 'SOLO TIENE TRADUCCIÓN PARA IDIOMA "ES"'
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1 
        ####
        if traduc3.empty != True:
            ##
            if "TRANSLATION" not in tipo_verificacion:
                tipo_verificacion.append("TRANSLATION")
            ##
            rows = dataframe_to_rows(traduc3,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "NO TIENE REGISTRADO NINGUNA TRADUCCIÓN"
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1 
        ####
        ##
        if des_nom_product.empty != True:
            ##
            tipo_verificacion.append("TRANSLATIONNAMEPRODUCT")
            ##
            rows = dataframe_to_rows(des_nom_product,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "LA DESCRIPCIÓN O EL NOMBRE DEL PRODUCTO EN TRADUCCIONES DE TEXTO ESTÁ VACIO."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1 
        ####
        
        ####
        if product_type.empty != True:
            ##
            tipo_verificacion.append("PRODUCTTYPE")
            ##
            rows = dataframe_to_rows(product_type,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "EL TIPO DE PRODUCTO EN GENERAL NO ESTÁ COMO 'Artículo'."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ####
        ####
        if product_subtype.empty != True:
            ##
            tipo_verificacion.append("PRODUCTSUBTYPE")
            ##
            rows = dataframe_to_rows(product_subtype,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "EL SUBTIPO DE PRODUCTO EN GENERAL NO ESTÁ COMO 'Producto'."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if service_type.empty != True:
            ##
            tipo_verificacion.append("SERVICETYPE")
            ##
            rows = dataframe_to_rows(service_type,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "EL TIPO DE SERVICIO DE PRODUCTO EN GENERAL NO ESTÁ COMO 'No especificado'."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if name_fields.empty != True:
            ##
            tipo_verificacion.append("NAMEFIELDS")
            ##
            rows = dataframe_to_rows(name_fields,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "ALGUNO(S) DE LOS CAMPOS DE 'NOMBRE DE PRODUCTO', 'NOMBRE DE BÚSQUEDA' O 'DESCRIPCIÓN' EN GENERAL, ESTÁ(N) VACIO(S)."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if storage_dimension.empty != True:
            ##
            tipo_verificacion.append("STORAGEDIMENSION")
            ##
            rows = dataframe_to_rows(storage_dimension,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "EL CAMPO DE 'GRUPO DE DIMENSIONES DE ALMACENAMIENTO' EN GENERAL ES DISTINTO A 'SiteWHL'."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if dimension_group_name.empty != True:
            ##
            tipo_verificacion.append("DIMENSIONGROUPNAME")
            ##
            rows = dataframe_to_rows(dimension_group_name,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "EL CAMPO DE 'GRUPO DE DIMENSIONES DE SEGUIMIENTO' EN GENERAL ES DISTINTO A 'Ninguno'."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if item_model.empty != True:
            ##
            tipo_verificacion.append("ITEMMODEL")
            ##
            rows = dataframe_to_rows(item_model,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "EL CAMPO DE 'GRUPO DE MODELOS DE ARTÍCULO' EN GENERAL ES DISTINTO A 'PROD'."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if purchase_unit_symbol.empty != True:
            ##
            tipo_verificacion.append("PURCHASEUNITSYMBOL")
            ##
            rows = dataframe_to_rows(purchase_unit_symbol,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "LA 'UNIDAD' EN PEDIDO DE COMPRA ES IGUAL A 'U.'."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if purchase_sales_tax_item.empty != True:
            ##
            tipo_verificacion.append("PURCHASESALESTAXITEM")
            ##
            rows = dataframe_to_rows(purchase_sales_tax_item,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "EL CAMPO DE 'GRUPO DE IMPUESTOS DE ARTÍCULOS' EN COMPRA ESTA VACIO."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if sales_sales_tax_item.empty != True:
            ##
            tipo_verificacion.append("SALESSALESTAXITEM")
            ##
            rows = dataframe_to_rows(sales_sales_tax_item,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "EL CAMPO DE 'GRUPO DE IMPUESTOS DE ARTÍCULOS' EN VENDER ESTA VACIO."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if purchase_price_automaticall.empty != True:
            ##
            tipo_verificacion.append("PURCHASEPRICEAUTOMATICALL")
            ##
            rows = dataframe_to_rows(purchase_price_automaticall,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "EL 'ULTIMO PRECIO DE COMRA' EN 'COMPRA' ESTÁ MARCADO COMO 'NO'."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if net_product.empty != True:
            ##
            tipo_verificacion.append("NETPRODUCT")
            ##
            rows = dataframe_to_rows(net_product,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "EL CAMPO DE PESO NETO ESTA VACIO."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if order_underde.empty != True:
            ##
            tipo_verificacion.append("ORDERUNDERDE")
            ##
            rows = dataframe_to_rows(order_underde,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "EL CAMPO DE 'ENTREGA INCOMPLETA' EN 'ADMINISTRAR INVENTARIO' NO ES IGUAL A '100.00."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if production_type.empty != True:
            ##
            tipo_verificacion.append("PRODUCTIONTYPE")
            ##
            rows = dataframe_to_rows(production_type,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "EL CAMPO DE 'TIPO DE PRODUCCIÓN' EN 'APLICAR INGENIERÍA' NO ES IGUAL A 'NINGUNO'."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if product_coverage.empty != True:
            ##
            tipo_verificacion.append("PRODUCTCOVERAGE")
            ##
            rows = dataframe_to_rows(product_coverage,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "EL CAMPO DE 'GRUPO DE COBERTURA' DE 'PLAN' ES DISTINTO A 'GRUPO'."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if product_group_id.empty != True:
            ##
            tipo_verificacion.append("PRODUCTGROUPID")
            ##
            rows = dataframe_to_rows(product_group_id,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "EL CAMPO DE 'GRUPO DE ARTÍCULOS' DE 'GESTIONAR COSTES' ES DISTINTO A 'GA001'."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if unit_cost_automatically.empty != True:
            ##
            tipo_verificacion.append("UNITCOSTAUTOMATICALLY")
            ##
            rows = dataframe_to_rows(unit_cost_automatically,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "'ULTIMO PRECIO DE COSTE' DE 'GESTIONAR COSTES' ESTÁ MARCADO COMO 'NO'."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if code_product.empty != True:
            ##
            tipo_verificacion.append("CODEPRODUCTVERIFY")
            ##
            rows = dataframe_to_rows(code_product,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "EL CÓDIGO INTERNACIONAL Y CÓDIGO PRODUCTO SUNAT SON DISTINTOS."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        ##
        ##
        if code_productEightDigits.empty != True:
            ##
            tipo_verificacion.append("CODEPRODUCTEIGHT")
            ##
            rows = dataframe_to_rows(code_productEightDigits,index=False)
            for r_i, row in enumerate(rows,1):
                if r_i != 1:
                    num_row = seed_cell_excel + count_f
                    sheet[f'B{num_row}'] = count_f+1
                    sheet[f'C{num_row}'] = "EL CÓDIGO PRODUCTO SUNAT NO TIENE 8 DÍGITOS. VERIFICAR QUE NO TENGA ESPACIOS VACIOS ANTES O DESPUES DEL CÓDIGO."
                    count_c = 0
                    for c_i, value in enumerate(row,1):
                        sheet[f'{colum_l[count_c]}{num_row}'] = str(value)
                        count_c = count_c+1
                    count_f = count_f+1
        
        # Obtener la ruta al directorio 'assets' desde 'modulo1.py'
        ruta_assets = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'assets'))

        # Construir la ruta al archivo Excel dentro de la carpeta 'assets'
        ruta_archivo = os.path.join(ruta_assets, 'ReporteVerificacionProductos.xlsx')

        data_sheet.save(ruta_archivo)
        
        return tipo_verificacion, ruta_archivo
        
            
        ####